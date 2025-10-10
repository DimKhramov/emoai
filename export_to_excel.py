#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для экспорта всех сообщений из базы данных в Excel файл
"""

import os
import sys
import sqlite3
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Добавляем путь к корневой папке проекта
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from config import DatabaseConfig

def export_messages_to_excel():
    """Экспортировать все сообщения в Excel файл"""
    
    print("📊 Экспорт сообщений в Excel...")
    print("=" * 50)
    
    try:
        # Подключение к базе данных
        conn = sqlite3.connect(DatabaseConfig.DB_PATH, timeout=10.0)
        cursor = conn.cursor()
        
        # SQL запрос для получения всех сообщений с информацией о пользователях
        query = """
        SELECT 
            u.telegram_id,
            u.username,
            m.role,
            m.message,
            m.created_at,
            DATE(m.created_at) as message_date,
            TIME(m.created_at) as message_time
        FROM messages m
        JOIN users u ON m.user_id = u.id
        ORDER BY m.created_at ASC
        """
        
        print("🔍 Получение данных из базы...")
        cursor.execute(query)
        messages = cursor.fetchall()
        
        if not messages:
            print("❌ Сообщения не найдены в базе данных")
            return
        
        print(f"✅ Найдено {len(messages)} сообщений")
        
        # Создаем DataFrame
        df = pd.DataFrame(messages, columns=[
            'Telegram_ID', 'Username', 'Role', 'Message', 
            'Full_DateTime', 'Date', 'Time'
        ])
        
        # Обработка данных
        df['Username'] = df['Username'].fillna('Неизвестно')
        df['Role_RU'] = df['Role'].map({'user': 'Пользователь', 'assistant': 'Бот'})
        df['Message_Length'] = df['Message'].str.len()
        
        # Создаем Excel файл
        filename = f"chat_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        
        print(f"📝 Создание Excel файла: {filename}")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Основной лист с сообщениями
            df_export = df[['Telegram_ID', 'Username', 'Role_RU', 'Message', 'Date', 'Time', 'Message_Length']]
            df_export.to_excel(writer, sheet_name='Все сообщения', index=False)
            
            # Статистика по пользователям
            user_stats = df.groupby(['Telegram_ID', 'Username']).agg({
                'Message': 'count',
                'Message_Length': 'mean',
                'Date': ['min', 'max']
            }).round(2)
            user_stats.columns = ['Всего_сообщений', 'Средняя_длина', 'Первое_сообщение', 'Последнее_сообщение']
            user_stats.reset_index().to_excel(writer, sheet_name='Статистика пользователей', index=False)
            
            # Статистика по дням
            daily_stats = df.groupby(['Date', 'Role_RU']).size().unstack(fill_value=0)
            daily_stats['Всего'] = daily_stats.sum(axis=1)
            daily_stats.to_excel(writer, sheet_name='Статистика по дням')
            
            # Форматирование основного листа
            workbook = writer.book
            worksheet = writer.sheets['Все сообщения']
            
            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            user_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
            bot_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # Форматирование заголовков
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Ограничиваем максимальную ширину
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Цветовое кодирование по ролям
            for row in range(2, len(df_export) + 2):
                role = worksheet[f'C{row}'].value
                if role == 'Пользователь':
                    for col in range(1, 8):
                        worksheet.cell(row=row, column=col).fill = user_fill
                elif role == 'Бот':
                    for col in range(1, 8):
                        worksheet.cell(row=row, column=col).fill = bot_fill
        
        # Статистика
        total_messages = len(messages)
        user_messages = len(df[df['Role'] == 'user'])
        bot_messages = len(df[df['Role'] == 'assistant'])
        unique_users = df['Telegram_ID'].nunique()
        date_range = f"{df['Date'].min()} - {df['Date'].max()}"
        
        print("\n📊 Статистика экспорта:")
        print(f"   📝 Всего сообщений: {total_messages}")
        print(f"   👤 От пользователей: {user_messages}")
        print(f"   🤖 От бота: {bot_messages}")
        print(f"   👥 Уникальных пользователей: {unique_users}")
        print(f"   📅 Период: {date_range}")
        print(f"\n✅ Файл сохранен: {filepath}")
        
        conn.close()
        return filepath
        
    except Exception as e:
        print(f"❌ Ошибка при экспорте: {e}")
        return None

def create_summary_report():
    """Создать краткий отчет по экспорту"""
    
    try:
        conn = sqlite3.connect(DatabaseConfig.DB_PATH, timeout=10.0)
        cursor = conn.cursor()
        
        # Общая статистика
        cursor.execute("SELECT COUNT(*) FROM messages")
        total_messages = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM users")
        total_users = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM messages WHERE role = 'user'")
        user_messages = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM messages WHERE role = 'assistant'")
        bot_messages = cursor.fetchone()[0]
        
        # Топ активных пользователей
        cursor.execute("""
        SELECT u.username, u.telegram_id, COUNT(m.id) as msg_count
        FROM users u
        JOIN messages m ON u.id = m.user_id
        GROUP BY u.id
        ORDER BY msg_count DESC
        LIMIT 5
        """)
        top_users = cursor.fetchall()
        
        print("\n" + "="*60)
        print("📋 КРАТКИЙ ОТЧЕТ ПО ЭКСПОРТУ")
        print("="*60)
        print(f"📊 Общая статистика:")
        print(f"   • Всего сообщений: {total_messages}")
        print(f"   • Всего пользователей: {total_users}")
        print(f"   • Сообщений от пользователей: {user_messages}")
        print(f"   • Сообщений от бота: {bot_messages}")
        
        print(f"\n👑 Топ-5 активных пользователей:")
        for i, (username, telegram_id, msg_count) in enumerate(top_users, 1):
            print(f"   {i}. @{username or 'неизвестно'} (ID: {telegram_id}) - {msg_count} сообщений")
        
        conn.close()
        
    except Exception as e:
        print(f"❌ Ошибка при создании отчета: {e}")

if __name__ == "__main__":
    print("🚀 Запуск экспорта сообщений в Excel")
    
    # Экспорт в Excel
    excel_file = export_messages_to_excel()
    
    if excel_file:
        # Краткий отчет
        create_summary_report()
        
        print(f"\n🎉 Экспорт завершен успешно!")
        print(f"📁 Файл: {excel_file}")
    else:
        print("❌ Экспорт не удался")